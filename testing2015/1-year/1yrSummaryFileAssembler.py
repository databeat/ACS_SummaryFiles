#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec 16 11:49:29 2016

@author: jonathanortiz
"""

import os
import csv
import xlrd
import openpyxl as oxl


#################
""" Parameters """
# Provide a list of sequences from ACS_1yr_Seq_Table_Number_Lookup.xls
# Either provide your own list of sequences or simply comment out topic areas you do not need
'''
sequences = [
             1, # Unweighted Count
             2, 3, # Age-Sex
             4, 5, # Race
             6, # Hispanic Origin
             7, 8, 9, # Ancestry
             10, 11, 12, 13, 14, # Foreign Birth
             15, 16, 17, 18, # Place of Birth - Native
#             19, 20, 21, 22, 23, 24, 25, 26, 27, 28, # Residence Last Year - Migration
#             29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, # Journey to Work
#             45, # Children - Relationship
#             46, # Grand(Persons) - Age of HH Members
#             47, 48, # Households - Families
#             49, 50, # Marital Status
#             51, # Fertility
             52, 53, 54, # School Enrollment
             55, 56, 57, # Educational Attainment
#             58, 59, 60, # Language
             61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, # Poverty
#             73, 74, 75, 76, # Disability
             77, 78, 79, 80, 81, 82, 83, 84, # Income
             85, 86, 87, 88, 89, 90, # Earnings
#             91, 92, 93, # Veteran Status
#             94, # Transfer Programs (Food Stamps/SNAP)
             95, 96, 97, 98, 99, 100, 101, 102, # Employment Status
#             103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, # Industry-Occupation-Class of Worker
             138, 139, 140, 141, 142, 143, 144, 145, 146, 147, 148, # Housing
#             149, # Group Quarters
             150, 151, 152, 153, 154, 155, 156, 157, 158, 159, # Health Insurance
             160, # Computer and Internet Usage
             161, # Quality Measures
#             162, 163, 164, 165 # Imputations
             ]
'''
sequences = [1, 2, 66, 67]

# Provide a list of geographies from 1_year_Mini_Geo.xlsx
# Note: these codes must be lowercase to match those in 1_year_Mini_Geo.xlsx
'''
geos = ["al", "ak", "az", "ar", "ca", "co", "ct", "de", "fl", "ga", "hi", "id",
       "il", "in", "ia", "ks", "ky", "la", "me", "md", "ma", "mi", "mn", "ms",
       "mo", "mt", "ne", "nv", "nh", "nj", "nm", "ny", "nc", "nd", "oh", "ok",
       "or", "pa", "ri", "sc", "sd", "tn", "tx", "us", "ut", "vt", "va", "wa",
       "wv", "wi", "wy"]
'''
geos = ["ca", "dc", "wy"]

# Provide a choice of estimates, margins-of-error, or both
# Note: these codes must by uppercase to match those in the Sequence files
measures = ["E", "M"]

# Note: Do not change topics dictionary. It is required to assemble the output directories
topics = {1: "Unweighted Count", 2: "Age-Sex", 3: "Age-Sex", 4: "Race", 5: "Race", 6: "Hispanic Origin", 7: "Ancestry", 8: "Ancestry", 9: "Ancestry", 10: "Foreign Birth", 11: "Foreign Birth", 12: "Foreign Birth", 13: "Foreign Birth", 14: "Foreign Birth", 15: "Place of Birth - Native", 16: "Place of Birth - Native", 17: "Place of Birth - Native", 18: "Place of Birth - Native", 19: "Residence Last Year - Migration", 20: "Residence Last Year - Migration", 21: "Residence Last Year - Migration", 22: "Residence Last Year - Migration", 23: "Residence Last Year - Migration", 24: "Residence Last Year - Migration", 25: "Residence Last Year - Migration", 26: "Residence Last Year - Migration", 27: "Residence Last Year - Migration", 28: "Residence Last Year - Migration", 29: "Journey to Work", 30: "Journey to Work", 31: "Journey to Work", 32: "Journey to Work", 33: "Journey to Work", 34: "Journey to Work", 35: "Journey to Work", 36: "Journey to Work", 37: "Journey to Work", 38: "Journey to Work", 39: "Journey to Work", 40: "Journey to Work", 41: "Journey to Work", 42: "Journey to Work", 43: "Journey to Work", 44: "Journey to Work", 45: "Children - Relationship", 46: "Grand(Persons) - Age of HH Members", 47: "Households - Families", 48: "Households - Families", 49: "Marital Status", 50: "Marital Status", 51: "Fertility", 52: "School Enrollment", 53: "School Enrollment", 54: "School Enrollment", 55: "Educational Attainment", 56: "Educational Attainment", 57: "Educational Attainment", 58: "Language", 59: "Language", 60: "Language", 61: "Poverty", 62: "Poverty", 63: "Poverty", 64: "Poverty", 65: "Poverty", 66: "Poverty", 67: "Poverty", 68: "Poverty", 69: "Poverty", 70: "Poverty", 71: "Poverty", 72: "Poverty", 73: "Disability", 74: "Disability", 75: "Disability", 76: "Disability", 77: "Income", 78: "Income", 79: "Income", 80: "Income", 81: "Income", 82: "Income", 83: "Income", 84: "Income", 85: "Earnings", 86: "Earnings", 87: "Earnings", 88: "Earnings", 89: "Earnings", 90: "Earnings", 91: "Veteran Status", 92: "Veteran Status", 93: "Veteran Status", 94: "Transfer Programs (Food Stamps/SNAP)", 95: "Employment Status", 96: "Employment Status", 97: "Employment Status", 98: "Employment Status", 99: "Employment Status", 100: "Employment Status", 101: "Employment Status", 102: "Employment Status", 103: "Industry-Occupation-Class of Worker", 104: "Industry-Occupation-Class of Worker", 105: "Industry-Occupation-Class of Worker", 106: "Industry-Occupation-Class of Worker", 107: "Industry-Occupation-Class of Worker", 108: "Industry-Occupation-Class of Worker", 109: "Industry-Occupation-Class of Worker", 110: "Industry-Occupation-Class of Worker", 111: "Industry-Occupation-Class of Worker", 112: "Industry-Occupation-Class of Worker", 113: "Industry-Occupation-Class of Worker", 114: "Industry-Occupation-Class of Worker", 115: "Industry-Occupation-Class of Worker", 116: "Industry-Occupation-Class of Worker", 117: "Industry-Occupation-Class of Worker", 118: "Industry-Occupation-Class of Worker", 119: "Industry-Occupation-Class of Worker", 120: "Industry-Occupation-Class of Worker", 121: "Industry-Occupation-Class of Worker", 122: "Industry-Occupation-Class of Worker", 123: "Industry-Occupation-Class of Worker", 124: "Industry-Occupation-Class of Worker", 125: "Industry-Occupation-Class of Worker", 126: "Industry-Occupation-Class of Worker", 127: "Industry-Occupation-Class of Worker", 128: "Industry-Occupation-Class of Worker", 129: "Industry-Occupation-Class of Worker", 130: "Industry-Occupation-Class of Worker", 131: "Industry-Occupation-Class of Worker", 132: "Industry-Occupation-Class of Worker", 133: "Industry-Occupation-Class of Worker", 134: "Industry-Occupation-Class of Worker", 135: "Industry-Occupation-Class of Worker", 136: "Industry-Occupation-Class of Worker", 137: "Industry-Occupation-Class of Worker", 138: "Housing", 139: "Housing", 140: "Housing", 141: "Housing", 142: "Housing", 143: "Housing", 144: "Housing", 145: "Housing", 146: "Housing", 147: "Housing", 148: "Housing", 149: "Group Quarters", 150: "Health Insurance", 151: "Health Insurance", 152: "Health Insurance", 153: "Health Insurance", 154: "Health Insurance", 155: "Health Insurance", 156: "Health Insurance", 157: "Health Insurance", 158: "Health Insurance", 159: "Health Insurance", 160: "Computer and Internet Usage", 161: "Quality Measures", 162: "Imputations", 163: "Imputations", 164: "Imputations", 165: "Imputations"}

#################
""" Functions """
def getTopicAreas(seqs):
    topxSet = set()
    for seq in seqs:
        topxSet.add(topics[seq])
    topx = list(topxSet)
    return topx


def ensureDirs(topicAreas):
    for topic in topicAreas:
        d = os.getcwd() + "/output/" + topic + "/"
        if not os.path.exists(d):
            os.makedirs(d)


def getSeqHeader(seq, EorM):
    header = []
    inFileName = os.getcwd() + "/data/1_year_Summary_FileTemplates/Seq" + str(seq) + ".xls"
    seqFile = xlrd.open_workbook(inFileName)
    sheet = seqFile.sheet_by_name(EorM)
    numCols = sheet.ncols
    for col in range(numCols):
        header.append(sheet.cell_value(1, col))
    header.insert(6, "GEOID")
    header.insert(7, "GEONAME")
    return header


def getSummaryData(seq, EorM, geo):
    data = []
    number = str(seq).zfill(4)
    inFileName = os.getcwd() + "/data/1_year_entire_sf/" + EorM.lower() + "20141" + geo + number + "000.txt"
    with open(inFileName, 'r', newline='') as inFile:
        reader = csv.reader(inFile, delimiter=',', quotechar='"')
        for row in reader:
            data.append(row)
    return data


def getGeoCodes(geo):
    geoCodes = []
    inFileName = os.getcwd() + "/documentation/geography/1_year_Mini_Geo.xlsx"
    wb = oxl.load_workbook(filename = inFileName)
    sheet = wb[geo]
    rows = sheet.max_row
    for i in range(2,rows+1):
        geoid = sheet.cell(row=i, column=2).value
        geoname = sheet.cell(row=i, column=3).value
        geoCodes.append([geoid, geoname])
    return geoCodes

    
def assemble(seq, EorM, geo, header, geoCodes, data):
    topic = topics[seq]
    number = str(seq).zfill(3)
    outFileName = os.getcwd() + "/output/" + topic + "/" + geo.upper() + "_" + topic + "_" + number + EorM.lower() + ".csv"    
    with open(outFileName, 'w+', newline='') as outFile:
        writer = csv.writer(outFile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(header)
        for row in data:
            index = data.index(row)
            g = geoCodes[index]
            row.insert(6, g[0])
            row.insert(7, g[1])
            writer.writerow(row)


############
""" Main """

topicAreas = getTopicAreas(sequences)

ensureDirs(topicAreas)

for geo in geos:
    geoCodes = getGeoCodes(geo)
    for seq in sequences:
        for EorM in measures:
            header = getSeqHeader(seq, EorM)
            data = getSummaryData(seq, EorM, geo)
            assemble(seq, EorM, geo, header, geoCodes, data)
